import sys
import openpyxl
import os
os.chdir("D:\\revere\\office\\uploads");
print("Output from Python")
mylist = sys.argv[1].split(',')
destination = sys.argv[2]
template = sys.argv[3]
vendorColumn = 25
if(template == "hoardingsMasterSheet"):
    vendorColumn = 25
elif(template == "bqsMasterSheet"):
    vendorColumn = 27
mainWb = openpyxl.load_workbook(destination)
mainsheet = mainWb.active
for row in mainsheet.rows:
    l = []
    for cell in row:
        new = []
        if(cell.font.bold == True and cell.value != None):
            new.append(str(cell.column))
            new.append(str(cell.row))
            new.append(str(cell.value).lower().strip())
        if(new):
            l.append(new)
for a in mylist:
    max_row = mainsheet.max_row
    max_column = len(mainsheet['1'])
    wb = openpyxl.load_workbook(a,data_only=True)
    sheet = wb.active
    print(a)
    d = 0
    for row in sheet.rows:
        ## to stop after checking 5 rows in a sheet
        if(d<5):
            d = d+1
            b = 0
            for cell in row:
                if(cell.font.bold == True and cell.value != None):
                    b = b+1
            ## to check if no of bolded cell is grater than 2
            if(b>2):
                for cell in row:
                    if(cell.font.bold == True and cell.value != None):
                        headingVal = cell.value.lower().strip()
                        ## find if heading has xxx for '*' and 'x' type of notations
                        if(headingVal.find("xxx") != -1):
                            c1 = 0
                            c2 = 0
                        ##  identify which
                            if(headingVal.find("size") != -1 and template == "hoardingsMasterSheet"):
                                c1 = 6
                                c2 = 7
                            elif(template == "bqsMasterSheet"):
                                if(headingVal.find("top") != -1):
                                    c1 = 10
                                    c2 = 12
                                elif(headingVal.find("side") != -1):
                                    c1 = 13
                                    c2 = 15
                                elif(headingVal.find("back") != -1):
                                    c1 = 17
                                    c2 = 19
                                elif(headingVal.find("mupi") != -1):
                                    c1 = 21
                                    c2 = 23
                            if(c1 != 0):
                                r = cell.row
                                c = cell.column
                                print("row-" + str(r) + ",column-" + str(c) + ",value-" + str(headingVal) + " matches -" + str("xxx size") )
                                for i in range(r,sheet.max_row):
                                    if(sheet.cell(row = i+1,column = c).value != None):
                                        v = sheet.cell(row = i+1,column = c).value.lower()
                                        if(v.find('x') != -1):
                                            v = v.split('x')
                                        elif(v.find('*') != -1):
                                            v = v.split('*')
                                        if(len(v) == 2):
                                            mainsheet.cell(row = max_row + i + 1 -r,column = c1).value = float(str(v[0]).strip().strip("\'").strip("\"").replace("\'","."))
                                            mainsheet.cell(row = max_row + i + 1 -r,column = c2).value = float(str(v[1]).strip().strip("\'").strip("\"").replace("\'","."))
                        else:
                            matching = [s for s in l if cell.value.lower().strip() in s]
                            r = cell.row
                            c = cell.column
                            v = cell.value
                            if(matching):
                                if(matching[0][0] != "1"):
                                    if(len(matching)>1):
                                        matching = [matching[1]]
        ##                          copy all following rows data in column to matching column in main sheet                    
                                    print("row-" + str(r) + ",column-" + str(c) + ",value-" + str(v) + " matches -" + str(matching) )
                                    for i in range(r,sheet.max_row):
                                        if(mainsheet.cell(row = max_row + i + 1 -r,column = 1).value == None):
                                            mainsheet.cell(row = max_row + i + 1 -r,column = 1).value = max_row + i + 1 -r -3
        ##                                  get column by finding vendor name in headings  
                                            mainsheet.cell(row = max_row + i + 1 -r,column = vendorColumn).value = ".".join(a[14:].split('.')[0:-1])
                                        valueToBeUpdated = sheet.cell(row = i+1,column = c).value
                                        if(matching[0][2] == 'h\'' or matching[0][2] == 'w\'' or matching[0][2] == 'tpsw\'' or matching[0][2] == 'tpsh\'' or matching[0][2] == 'spsw\'' or matching[0][2] == 'spsh\'' or matching[0][2] == 'bdw\'' or matching[0][2] == 'bdh\'' or matching[0][2] == 'mpw\'' or matching[0][2] == 'mph\'' ):
                                            try:
                                                valueToBeUpdated = float(str(valueToBeUpdated).strip().strip("\'").strip("\"").replace("\'","."))
                                            except ValueError:
                                                valueToBeUpdated = sheet.cell(row = i+1,column = c).value
                                        mainsheet.cell(row = max_row + i + 1 -r,column = int(matching[0][0])).value = valueToBeUpdated
                            else:
        ##                      create a new column in mainsheet and copy following rows data to new column in main sheet       
                                print("No match Found -- row-" + str(r) + ",column-" + str(c) + ",value-" + str(v))
                print();
print("print outside loop before saving")
mainWb.save(destination)
print("file saved to " + destination)
